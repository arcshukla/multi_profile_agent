"""
owner.py
--------
Owner portal routes.

All routes require the caller to be authenticated as role=owner or admin.
The profile slug is always taken from the session (never a URL param) so an
owner can only ever touch their own profile.

An admin visiting /owner/* is served the same view — slug comes from their
session["slug"] if set, otherwise they are redirected to /admin via require_owner.
"""

import io
import time
from datetime import date
from pathlib import Path

from fastapi import APIRouter, BackgroundTasks, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from app.core.config import TEMPLATES_DIR, settings
from app.core.constants import ALLOWED_DOC_EXTENSIONS, MAX_FILE_SIZE_PDF, MAX_FILE_SIZE_OTHER, MAX_DOCS_PER_PROFILE
from app.core.logging_config import get_logger
from app.auth.dependencies import require_owner
from app.services.billing_service import billing_service
from app.services.index_service import index_service
from app.services.preferences_service import preferences_service
from app.services.profile_service import profile_service
from app.services.prompt_service import prompt_service
from app.services.token_service import token_service
from app.services.user_service import user_service
from app.storage.file_storage import ProfileFileStorage
from app.rag.default_prompts import REQUIRED_PLACEHOLDERS, LOCKED_PROMPT_SUFFIXES
from app.utils.template_utils import render, htmx_ok, htmx_err
from app.utils.file_utils import sanitize_css

logger    = get_logger(__name__)
router    = APIRouter(prefix="/owner", include_in_schema=False)
templates = Jinja2Templates(directory=str(TEMPLATES_DIR))
templates.env.globals["support_email"] = settings.SUPPORT_EMAIL


def _r(request: Request, template: str, ctx: dict = None):
    return render(templates, request, template, ctx)


def _slug(user: dict) -> str:
    """Extract the profile slug from the session user dict."""
    return user.get("slug") or ""


# ── Dashboard ──────────────────────────────────────────────────────────────────

@router.get("/dashboard", response_class=HTMLResponse)
def dashboard(request: Request, user: dict = Depends(require_owner)):
    slug = _slug(user)
    return _r(request, "owner/dashboard.html", {
        "user":           user,
        "profile":        profile_service.get_profile(slug),
        "status":         index_service.get_status(slug),
        "usage":          token_service.get_all().get(slug, {}),
        "billing_entry":  billing_service.get_entry(slug),
        "billing_status": billing_service.get_status(slug),
    })


# ── Docs ───────────────────────────────────────────────────────────────────────

@router.get("/docs", response_class=HTMLResponse)
def docs_page(request: Request, user: dict = Depends(require_owner)):
    slug = _slug(user)
    fs   = ProfileFileStorage(slug)
    docs = fs.list_documents()
    return _r(request, "owner/docs.html", {
        "user":      user,
        "slug":      slug,
        "documents": [{"name": d.name, "size_kb": round(d.stat().st_size / 1024, 1)} for d in docs],
    })


@router.post("/docs/upload", response_class=HTMLResponse)
async def docs_upload(request: Request, file: UploadFile = File(...), user: dict = Depends(require_owner)):
    slug = _slug(user)
    fs   = ProfileFileStorage(slug)

    existing = fs.list_documents()
    if len(existing) >= MAX_DOCS_PER_PROFILE:
        return RedirectResponse(url="/owner/docs?upload_error=max_files", status_code=303)

    data     = await file.read()
    ext      = Path(file.filename).suffix.lower()
    max_size = MAX_FILE_SIZE_PDF if ext == ".pdf" else MAX_FILE_SIZE_OTHER
    limit_mb = max_size // (1024 * 1024)
    if len(data) > max_size:
        return RedirectResponse(
            url=f"/owner/docs?upload_error=too_large&ext={ext.lstrip('.')}&limit={limit_mb}",
            status_code=303,
        )

    try:
        fs.save_document(file.filename, data)
        logger.info("Owner upload: slug=%s file=%s size=%d", slug, file.filename, len(data))
    except ValueError as e:
        return RedirectResponse(url=f"/owner/docs?upload_error=invalid&msg={e}", status_code=303)

    return RedirectResponse(url="/owner/docs?reindex=1", status_code=303)


@router.post("/docs/delete/{filename}", response_class=HTMLResponse)
async def docs_delete(filename: str, user: dict = Depends(require_owner)):
    ProfileFileStorage(_slug(user)).delete_document(filename)
    return RedirectResponse(url="/owner/docs?reindex=1", status_code=303)


@router.get("/docs/view/{filename}")
async def docs_view(filename: str, user: dict = Depends(require_owner)):
    """Serve an uploaded document for in-browser preview (PDF / TXT)."""
    slug      = _slug(user)
    fs        = ProfileFileStorage(slug)
    safe_name = Path(filename).name
    if not safe_name or safe_name != filename:
        raise HTTPException(status_code=400, detail="Invalid filename")
    path = fs.docs_dir / safe_name
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    if path.suffix.lower() not in ALLOWED_DOC_EXTENSIONS:
        raise HTTPException(status_code=403, detail="File type not viewable")
    media_map = {
        ".pdf": "application/pdf",
        ".txt": "text/plain",
        ".md":  "text/plain",
        ".csv": "text/plain",
    }
    ext = path.suffix.lower()
    if ext not in media_map:
        raise HTTPException(status_code=403, detail="File type not viewable")
    return FileResponse(str(path), media_type=media_map[ext])


# ── Appearance (header + CSS) ─────────────────────────────────────────────────

@router.get("/appearance", response_class=HTMLResponse)
def appearance_page(request: Request, user: dict = Depends(require_owner)):
    slug      = _slug(user)
    fs        = ProfileFileStorage(slug)
    has_photo = fs.has_photo()
    return _r(request, "owner/appearance.html", {
        "user":        user,
        "slug":        slug,
        "has_photo":   has_photo,
        "photo_ts":    int(fs.photo_path.stat().st_mtime) if has_photo else 0,
        "slides_data": fs.read_slides(),
        "profile_css": fs.read_css(),
    })


@router.post("/appearance/slides")
async def save_slides(request: Request, user: dict = Depends(require_owner)):
    form   = await request.form()
    slides = []
    i = 0
    while f"type_{i}" in form:
        slide_type = form.get(f"type_{i}", "standard")
        if slide_type == "quote":
            slides.append({
                "type":        "quote",
                "quote":       form.get(f"quote_{i}", "").strip(),
                "attribution": form.get(f"attribution_{i}", "").strip(),
            })
        else:
            slides.append({
                "type":     "standard",
                "title":    form.get(f"title_{i}", "").strip(),
                "subtitle": form.get(f"subtitle_{i}", "").strip(),
                "body":     form.get(f"body_{i}", "").strip(),
            })
        i += 1
    slides = [s for s in slides if any(v for k, v in s.items() if k != "type")][:5]
    ProfileFileStorage(_slug(user)).write_slides({"slides": slides})
    return htmx_ok("Slides saved.")


@router.post("/appearance/css")
async def save_css(content: str = Form(...), user: dict = Depends(require_owner)):
    safe_css, violations = sanitize_css(content)
    if violations:
        logger.warning("save_css: blocked dangerous CSS from slug=%s | patterns=%s", _slug(user), violations)
        msg = (
            "CSS not saved — blocked content detected: "
            + ", ".join(f"<code>{v}</code>" for v in violations)
            + ". Remove <code>url()</code>, <code>@import</code>, and <code>javascript:</code> references."
        )
        return htmx_err(msg)
    ProfileFileStorage(_slug(user)).write_css(safe_css)
    return htmx_ok("CSS saved.")


# ── Prompts ────────────────────────────────────────────────────────────────────

@router.get("/prompts", response_class=HTMLResponse)
def prompts_page(request: Request, user: dict = Depends(require_owner)):
    slug     = _slug(user)
    prompts, _ = prompt_service.get_prompts(slug)
    return _r(request, "owner/prompts.html", {
        "user":     user,
        "prompts":  prompts,
        "required": REQUIRED_PLACEHOLDERS,
        "locked":   LOCKED_PROMPT_SUFFIXES,
    })


@router.post("/prompts/{key}")
async def save_prompt(key: str, content: str = Form(...), user: dict = Depends(require_owner)):
    required = REQUIRED_PLACEHOLDERS.get(key, [])
    missing  = [p for p in required if p not in content]
    if missing:
        msg = "Missing required placeholder(s): " + ", ".join(f"<code>{m}</code>" for m in missing)
        return htmx_err(msg)
    ok = prompt_service.update_prompt(_slug(user), key, content)
    if not ok:
        return htmx_err("Unknown prompt key.")
    return htmx_ok("Saved.")


# ── Photo ──────────────────────────────────────────────────────────────────────

@router.post("/photo")
async def upload_photo(request: Request, file: UploadFile = File(...), user: dict = Depends(require_owner)):
    slug = _slug(user)
    data = await file.read()
    if not data:
        return htmx_err("No file data received. Please select a file and try again.")
    ProfileFileStorage(slug).save_photo(data)
    ts = int(time.time())
    return _r(request, "owner/partials/photo_updated.html", {"slug": slug, "ts": ts})


# ── Status (enable / disable only — no delete for owners) ─────────────────────

@router.post("/status")
async def toggle_status(request: Request, status: str, user: dict = Depends(require_owner)):
    if status not in ("enabled", "disabled"):
        return htmx_err("Invalid status.")
    profile_service.update_status(_slug(user), status)
    return _r(request, "owner/partials/status_button.html", {"status": status})


# ── Analytics ─────────────────────────────────────────────────────────────────

@router.get("/analytics", response_class=HTMLResponse)
def analytics_page(request: Request, user: dict = Depends(require_owner)):
    slug   = _slug(user)
    events = ProfileFileStorage(slug).read_chat_events(limit=200)
    return _r(request, "owner/analytics.html", {
        "user":   user,
        "events": events,
    })


@router.get("/analytics/download")
def analytics_download(user: dict = Depends(require_owner)):
    """Download all chat events as an Excel workbook."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed — run: uv pip install openpyxl")

    slug   = _slug(user)
    events = list(reversed(ProfileFileStorage(slug).read_chat_events(limit=100_000)))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chat History"

    header_fill = PatternFill("solid", fgColor="EFF6FF")
    headers = ["Timestamp (UTC)", "Session ID", "Question", "Answer", "Tokens", "Answered"]
    for col, h in enumerate(headers, 1):
        cell       = ws.cell(row=1, column=col, value=h)
        cell.font  = Font(bold=True)
        cell.fill  = header_fill

    wrap_top = Alignment(wrap_text=True, vertical="top")
    top      = Alignment(vertical="top")

    for row_idx, e in enumerate(events, 2):
        ws.cell(row=row_idx, column=1, value=e.get("ts", "")).alignment          = top
        ws.cell(row=row_idx, column=2, value=e.get("session_id", "")).alignment  = top
        ws.cell(row=row_idx, column=3, value=e.get("question", "")).alignment    = wrap_top
        ws.cell(row=row_idx, column=4, value=e.get("answer", "")).alignment      = wrap_top
        ws.cell(row=row_idx, column=5, value=e.get("tokens") or 0).alignment     = top
        ws.cell(row=row_idx, column=6, value="Yes" if e.get("was_answered") else "No").alignment = top

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 70
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"{slug}-chat-history-{date.today().isoformat()}.xlsx"
    logger.info("Analytics download: slug=%s events=%d file=%s", slug, len(events), filename)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── AI & Indexing ──────────────────────────────────────────────────────────────

@router.get("/ai", response_class=HTMLResponse)
def ai_page(request: Request, user: dict = Depends(require_owner)):
    slug = _slug(user)
    return _r(request, "owner/ai.html", {
        "user":    user,
        "usage":   token_service.get_all().get(slug, {}),
        "slug":    slug,
        "status":  index_service.get_status(slug),
        "history": index_service.get_history(slug, limit=20),
    })


@router.get("/tokens", response_class=HTMLResponse)
def tokens_redirect():
    """Backwards-compat redirect for old bookmarks."""
    return RedirectResponse(url="/owner/ai", status_code=301)


@router.post("/index")
async def owner_index(background_tasks: BackgroundTasks, user: dict = Depends(require_owner)):
    """Start indexing in background."""
    slug = _slug(user)
    if not index_service.is_indexing(slug):
        background_tasks.add_task(index_service.index_profile, slug, False)
    return RedirectResponse(url="/owner/ai?indexing=1", status_code=303)


@router.post("/index/force")
async def owner_force_index(background_tasks: BackgroundTasks, user: dict = Depends(require_owner)):
    """Force full re-index (wipes existing index)."""
    slug = _slug(user)
    if not index_service.is_indexing(slug):
        background_tasks.add_task(index_service.force_reindex, slug)
    return RedirectResponse(url="/owner/ai?indexing=1", status_code=303)


# ── Preferences ────────────────────────────────────────────────────────────────

@router.get("/preferences", response_class=HTMLResponse)
def preferences_page(request: Request, user: dict = Depends(require_owner)):
    slug  = _slug(user)
    prefs = preferences_service.get(slug)
    return _r(request, "owner/preferences.html", {
        "user":        user,
        "slug":        slug,
        "prefs":       prefs,
        "saved":       False,
        "error":       None,
        "active_page": "preferences",
    })


@router.post("/preferences", response_class=HTMLResponse)
async def preferences_save(
    request: Request,
    name:                    str = Form(""),
    notify_unanswered_email: str = Form(""),  # checkbox: "on" when checked, absent otherwise
    chat_history_limit:      str = Form(""),  # numeric string or empty
    user: dict = Depends(require_owner),
):
    slug  = _slug(user)
    error = None

    new_name = name.strip()
    if new_name:
        ok, err = user_service.update_name(user["email"], new_name)
        if not ok:
            error = err
        else:
            request.session["user"]["name"] = new_name

    # Parse and validate chat_history_limit
    from app.core.constants import CHAT_HISTORY_LIMIT_DEFAULT, CHAT_HISTORY_LIMIT_MIN, CHAT_HISTORY_LIMIT_MAX
    history_limit = CHAT_HISTORY_LIMIT_DEFAULT
    if chat_history_limit.strip():
        try:
            history_limit = max(CHAT_HISTORY_LIMIT_MIN, min(CHAT_HISTORY_LIMIT_MAX, int(chat_history_limit)))
        except ValueError:
            history_limit = CHAT_HISTORY_LIMIT_DEFAULT

    prefs = {
        "notify_unanswered_email": notify_unanswered_email == "on",
        "chat_history_limit":      history_limit,
    }
    if error is None:
        preferences_service.save(slug, prefs)

    return _r(request, "owner/preferences.html", {
        "user":        request.session.get("user", user),
        "slug":        slug,
        "prefs":       prefs,
        "saved":       error is None,
        "error":       error,
        "active_page": "preferences",
    })
